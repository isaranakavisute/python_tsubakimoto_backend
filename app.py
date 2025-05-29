from flask import Flask, jsonify, request
#from flaskext.mysql import MySQL
import mysql.connector.pooling
from datetime import datetime
import pytz
import openpyxl 
import asyncio
import requests


# mysql = MySQL()

app = Flask(__name__)

# app.config['MYSQL_DATABASE_HOST'] = '127.0.0.1';
# app.config['MYSQL_DATABASE_PORT'] = 3307;
# app.config['MYSQL_DATABASE_USER'] = 'isara';
# app.config['MYSQL_DATABASE_PASSWORD'] = '1234';
# app.config['MYSQL_DATABASE_DB'] = 'mydb';
# mysql.init_app(app)
# cursor = mysql.get_db().cursor()

# my_database = MySQL(app, prefix="my_database", host="localhost", user="isara", password="1234", db="mydb", port=8888,autocommit=True)
# connection = my_database.connect()
# cursor = connection.cursor()



connection_pool = mysql.connector.pooling.MySQLConnectionPool(
    pool_name="my_pool",
    pool_size=30,
    user="isara",
    password="1234",
    host="deploy-aws.com",
    port=3307,  #3306
    database="tsubakimoto" #akt1
)

@app.route('/')
def hello_world():
    return jsonify(message="This is  a test")

@app.route('/test_db', methods=['GET'])
def get_test_db():
    app.logger.info('/test_db')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()
    cursor.execute('select * from exchange_rate')
    data = cursor.fetchall()
    cursor.close()
    conn.close() 
    return jsonify(data)

@app.route('/master_data/upload', methods=['POST'])
async def get_masterdata_upload():
    app.logger.info('/master_data/upload')
    
    #request mysql connection from pool
    conn = connection_pool.get_connection()
    cursor = conn.cursor()
    
    # upload file
    file = request.files['file']
    fullfilename = file.filename
    onlyfilename = fullfilename.split('.')[0];
    onlyfilename = onlyfilename.replace(' ','_')
    onlyfilename = onlyfilename.replace('-','_')
    onlyfileext = fullfilename.split('.')[1];
    print(request.files);
    newpath = "uploaded_files/" + onlyfilename  + "_" + datetime.now(pytz.timezone('Asia/Bangkok')).strftime('%Y_%m_%d_%H_%M_%S') + "." + onlyfileext;
    app.logger.info("uploaded new file path : "+newpath)
    file.save(newpath)

    requests.post('http://deploy-aws.com:5000/master_history/add', data={"master_file_name":onlyfilename  + "_" + datetime.now(pytz.timezone('Asia/Bangkok')).strftime('%Y_%m_%d_%H_%M_%S') + "." + onlyfileext,"file_path":newpath})

    # parse file
    wb = openpyxl.load_workbook(newpath,data_only=True)
    ws = wb.active
    print('Total number of rows: '+str(ws.max_row)+'. And total number of columns: '+str(ws.max_column))
    for row in range(5, ws.max_row+1):
        sql="insert into master_tsubakimoto(category,part_no,previous_model_no,new_model_no,unit,manufacturer_suggested_retail_price,new_manufacturer_suggested_retail_price,conversion_to_ft,diff_for_cost,op_price,po_price_jpy_usd,po_price_currency,remark,thb_cost,gp,pricelist_name,multiplier,make_same_price_as_standard_price,new_make_same_price_as_standard_price,standard_price,diff,dist_pl_mull,dist_ex_rate,unit_price,new_unit_price,diff_unit_price,status,supplier_name,stock_reference,cutting_assembly,detail)";
        sql += " values (";
        for column in range(1, ws.max_column+1):
            val = ws.cell(row,column).value
            if val is str:
                val = val.replace('\n','')
                val = val.replace('\r','')
                val = val.replace('\t','')
            elif val is None or val == '#VALUE!':
                val = "";
            if column < ws.max_column:
                sql += "'"
                sql += str(val);
                sql += "',"
                if val == "":
                    print("", end=",")
                else:
                    print(val, end=",")
            else:
                sql += "'"
                sql += str(val)
                sql += "')"
                if val == "":
                    print("", end="")
                else:
                    print(val, end="")
        print()
        
        #print sql for reviewing
        print("sql="+sql);
        
        #run sql
        cursor.execute(sql)
        
        print()
        print()
    
    
    
    data = { 
        "status":"true",
        "upload_excel":
        {
        "result": "pass",
        "full uploaded file path": newpath
        }
        } 
    
    #commit changes to databse
    conn.commit()
    
    #return mysql connection to pool
    cursor.close()
    conn.close() 
    
    await asyncio.sleep(5)
    
    #return json response
    return jsonify(data)

@app.route('/master_data/listall', methods=['POST'])
def get_masterdata_listall():
    app.logger.info('/master_data/listall')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()
    cursor.execute('select * from master_tsubakimoto')
    data = cursor.fetchall()
    cursor.close()
    conn.close()
    return jsonify(data)

@app.route('/master_data/deleteall', methods=['POST'])
def get_masterdata_deleteall():
    app.logger.info('/master_data/deleteall')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()
    cursor.execute('delete from master_tsubakimoto')
    # data = cursor.fetchall()
    conn.commit()
    cursor.close()
    conn.close()
    data = {
        "status":"true",
        "delete all masterdata":
            {
                "result": "pass"
            }
    }
    return jsonify(data)

@app.route('/master_data/update', methods=['POST'])
def get_masterdata_update():
    app.logger.info('/master_data/update')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()

    sql = "update master_tsubakimoto set "

    if request.form.get('category') is not None:
        sql += ","
        sql += "category='"
        sql += request.form.get('category')
        sql += "'"

    if request.form.get('part_no') is not None:
        sql += ","
        sql += "part_no='"
        sql += request.form.get('part_no')
        sql += "'"

    if request.form.get('previous_model_no') is not None:
        sql += ","
        sql += "previous_model_no='"
        sql += request.form.get('previous_model_no')
        sql += "'"

    if request.form.get('new_model_no') is not None:
        sql += ","
        sql += "new_model_no='"
        sql += request.form.get('new_model_no')
        sql += "'"

    if request.form.get('new_model_no') is not None:
        sql += ","
        sql += "unit='"
        sql += request.form.get('new_model_no')
        sql += "'"

    if request.form.get('manufacturer_suggested_retail_price') is not None:
        sql += ","
        sql += "manufacturer_suggested_retail_price='"
        sql += request.form.get('manufacturer_suggested_retail_price')
        sql += "'"

    if request.form.get('new_manufacturer_suggested_retail_price') is not None:
        sql += ","
        sql += "new_manufacturer_suggested_retail_price='"
        sql += request.form.get('new_manufacturer_suggested_retail_price')
        sql += "'"

    if request.form.get('conversion_to_ft') is not None:
        sql += ","
        sql += "conversion_to_ft='"
        sql += request.form.get('conversion_to_ft')
        sql += "'"

    if request.form.get('diff_for_cost') is not None:
        sql += ","
        sql += "diff_for_cost='"
        sql += request.form.get('diff_for_cost')
        sql += "'"

    if request.form.get('op_price') is not None:
        sql += ","
        sql += "op_price='"
        sql += request.form.get('op_price')
        sql += "'"

    if request.form.get('po_price_jpy_usd') is not None:
        sql += ","
        sql += "po_price_jpy_usd='"
        sql += request.form.get('po_price_jpy_usd')
        sql += "'"

    if request.form.get('po_price_currency') is not None:
        sql += ","
        sql += "po_price_currency='"
        sql += request.form.get('po_price_currency')
        sql += "'"

    if request.form.get('remark') is not None:
        sql += ","
        sql += "remark='"
        sql += request.form.get('remark')
        sql += "'"

    if request.form.get('thb_cost') is not None:
        sql += ","
        sql += "thb_cost='"
        sql += request.form.get('thb_cost')
        sql += "'"

    if request.form.get('gp') is not None:
        sql += ","
        sql += "gp='"
        sql += request.form.get('gp')
        sql += "'"

    if request.form.get('pricelist_name') is not None:
        sql += ","
        sql += "pricelist_name='"
        sql += request.form.get('pricelist_name')
        sql += "'"

    if request.form.get('multiplier') is not None:
        sql += ","
        sql += "multiplier='"
        sql += request.form.get('multiplier')
        sql += "'"

    if request.form.get('make_same_price_as_standard_price') is not None:
        sql += ","
        sql += "make_same_price_as_standard_price='"
        sql += request.form.get('make_same_price_as_standard_price')
        sql += "'"

    if request.form.get('new_make_same_price_as_standard_price') is not None:
        sql += ","
        sql += "new_make_same_price_as_standard_price='"
        sql += request.form.get('new_make_same_price_as_standard_price')
        sql += "'"

    if request.form.get('standard_price') is not None:
        sql += ","
        sql += "standard_price='"
        sql += request.form.get('standard_price')
        sql += "'"

    if request.form.get('diff') is not None:
        sql += ","
        sql += "diff='"
        sql += request.form.get('diff')
        sql += "'"

    if request.form.get('dist_pl_mul') is not None:
        sql += ","
        sql += "dist_pl_mull='"
        sql += request.form.get('dist_pl_mul')
        sql += "'"

    if request.form.get('dist_ex_rate') is not None:
        sql += ","
        sql += "dist_ex_rate='"
        sql += request.form.get('dist_ex_rate')
        sql += "'"

    if request.form.get('unit_price') is not None:
        sql += ","
        sql += "unit_price='"
        sql += request.form.get('unit_price')
        sql += "'"

    if request.form.get('new_unit_price') is not None:
        sql += ","
        sql += "new_unit_price='"
        sql += request.form.get('new_unit_price')
        sql += "'"

    if request.form.get('diff_unit_price') is not None:
        sql += ","
        sql += "diff_unit_price='"
        sql += request.form.get('diff_unit_price')
        sql += "'"

    if request.form.get('status') is not None:
        sql += ","
        sql += "status='"
        sql += request.form.get('status')
        sql += "'"

    if request.form.get('supplier_name') is not None:
        sql += ","
        sql += "supplier_name='"
        sql += request.form.get('supplier_name')
        sql += "'"

    if request.form.get('stock_reference') is not None:
        sql += ","
        sql += "stock_reference='"
        sql += request.form.get('stock_reference')
        sql += "'"

    if request.form.get('cutting_assembly') is not None:
        sql += ","
        sql += "cutting_assembly='"
        sql += request.form.get('cutting_assembly')
        sql += "'"

    if request.form.get('detail') is not None:
        sql += ","
        sql += "detail='"
        sql += request.form.get('detail')
        sql += "'"

    sql += " where Id="
    sql += request.form.get('Id')

    sql = sql.replace('update master_tsubakimoto set ,','update master_tsubakimoto set ')
    print('sql='+sql)

    cursor.execute(sql)
    conn.commit()

    cursor.close()
    conn.close()

    data = {
        "status":"true",
        "update masterdata":
            {
                "result": "pass"
            }
    }
    return jsonify(data)

@app.route('/master_data/add', methods=['POST'])
def get_masterdata_add():
    app.logger.info('/master_data/update')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()

    sql = "insert into master_tsubakimoto";
    sql += "(category,";
    sql += "part_no,";
    sql += "previous_model_no,";
    sql += "new_model_no,";
    sql += "unit,";
    sql += "manufacturer_suggested_retail_price,";
    sql += "new_manufacturer_suggested_retail_price,";
    sql += "conversion_to_ft,";
    sql += "diff_for_cost,";
    sql += "op_price,";
    sql += "po_price_jpy_usd,";
    sql += "po_price_currency,";
    sql += "remark,";
    sql += "thb_cost,";
    sql += "gp,";
    sql += "pricelist_name,";
    sql += "multiplier,";
    sql += "make_same_price_as_standard_price,";
    sql += "new_make_same_price_as_standard_price,";
    sql += "standard_price,";
    sql += "diff,";
    sql += "dist_pl_mull,";
    sql += "dist_ex_rate,";
    sql += "unit_price,";
    sql += "new_unit_price,";
    sql += "diff_unit_price,";
    sql += "status,";
    sql += "supplier_name,";
    sql += "stock_reference,";
    sql += "cutting_assembly,";
    sql += "detail)";
    sql += " values (";

    sql += "'"
    if request.form.get('category') is not None:
        sql += request.form.get('category')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('part_no') is not None:
     sql += request.form.get('part_no')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('previous_model_no') is not None:
        sql += request.form.get('previous_model_no')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('new_model_no') is not None:
        sql += request.form.get('new_model_no')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('unit') is not None:
        sql += request.form.get('unit')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('manufacturer_suggested_retail_price') is not None:
        sql += request.form.get('manufacturer_suggested_retail_price')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('new_manufacturer_suggested_retail_price') is not None:
        sql += request.form.get('new_manufacturer_suggested_retail_price')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('conversion_to_ft') is not None:
        sql += request.form.get('conversion_to_ft')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('diff_for_cost') is not None:
        sql += request.form.get('diff_for_cost')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('op_price') is not None:
        sql += request.form.get('op_price')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('po_price_jpy_usd') is not None:
        sql += request.form.get('po_price_jpy_usd')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('po_price_currency') is not None:
        sql += request.form.get('po_price_currency')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('remark') is not None:
        sql += request.form.get('remark')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('thb_cost') is not None:
        sql += request.form.get('thb_cost')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('gp') is not None:
        sql += request.form.get('gp')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('pricelist_name') is not None:
        sql += request.form.get('pricelist_name')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('multiplier') is not None:
     sql += request.form.get('multiplier')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('make_same_price_as_standard_price') is not None:
        sql += request.form.get('make_same_price_as_standard_price')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('new_make_same_price_as_standard_price') is not None:
        sql += request.form.get('new_make_same_price_as_standard_price')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('standard_price') is not None:
        sql += request.form.get('standard_price')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('diff') is not None:
        sql += request.form.get('diff')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('dist_pl_mull') is not None:
        sql += request.form.get('dist_pl_mull')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('dist_ex_rate') is not None:
        sql += request.form.get('dist_ex_rate')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('unit_price') is not None:
        sql += request.form.get('unit_price')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('new_unit_price') is not None:
        sql += request.form.get('new_unit_price')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('diff_unit_price') is not None:
        sql += request.form.get('diff_unit_price')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('status') is not None:
        sql += request.form.get('status')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('supplier_name') is not None:
        sql += request.form.get('supplier_name')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('stock_reference') is not None:
        sql += request.form.get('stock_reference')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('cutting_assembly') is not None:
        sql += request.form.get('cutting_assembly')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('detail') is not None:
        sql += request.form.get('detail')
    sql += "'"

    sql += ")";

    print('sql='+sql)

    cursor.execute(sql)
    conn.commit()

    cursor.close()
    conn.close()

    data = {
        "status":"true",
        "add masterdata":
            {
                "result": "pass"
            }
    }
    return jsonify(data)

@app.route('/master_data/delete', methods=['POST'])
def get_masterdata_delete():
    app.logger.info('/master_data/delete')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()

    sql = "delete from master_tsubakimoto"
    sql += " where Id="
    sql += request.form.get('Id')

    print('sql='+sql)

    cursor.execute(sql)
    conn.commit()

    cursor.close()
    conn.close()

    data = {
        "status":"true",
        "delete masterdata with id":
            {
                "result": "pass"
            }
    }
    return jsonify(data)

@app.route('/master_formula/upload', methods=['POST'])
async def get_masterformula_upload():
    app.logger.info('/master_formula/upload')

    #request mysql connection from pool
    conn = connection_pool.get_connection()
    cursor = conn.cursor()

    # upload file
    file = request.files['file']
    fullfilename = file.filename
    onlyfilename = fullfilename.split('.')[0];
    onlyfilename = onlyfilename.replace(' ','_')
    onlyfilename = onlyfilename.replace('-','_')
    onlyfileext = fullfilename.split('.')[1];
    print(request.files);
    newpath = "uploaded_files/" + onlyfilename  + "_" + datetime.now(pytz.timezone('Asia/Bangkok')).strftime('%Y_%m_%d_%H_%M_%S') + "." + onlyfileext;
    app.logger.info("uploaded new file path : "+newpath)
    file.save(newpath)

    # parse file
    wb = openpyxl.load_workbook(newpath,data_only=False)
    ws = wb.active
    print('Total number of rows: '+str(ws.max_row)+'. And total number of columns: '+str(ws.max_column))
    for row in range(5, ws.max_row+1):
        sql="insert into master_tsubakimoto_formula(category,part_no,previous_model_no,new_model_no,unit,manufacturer_suggested_retail_price,new_manufacturer_suggested_retail_price,conversion_to_ft,diff_for_cost,op_price,po_price_jpy_usd,po_price_currency,remark,thb_cost,gp,pricelist_name,multiplier,make_same_price_as_standard_price,new_make_same_price_as_standard_price,standard_price,diff,dist_pl_mull,dist_ex_rate,unit_price,new_unit_price,diff_unit_price,status,supplier_name,stock_reference,cutting_assembly,detail)";
        sql += " values (";
        for column in range(1, ws.max_column+1):
            val = ws.cell(row,column).value
            if val is str:
                val = val.replace('\n','')
                val = val.replace('\r','')
                val = val.replace('\t','')
            elif val is None or val == '#VALUE!':
                val = "";
            if column < ws.max_column:
                sql += "'"
                sql += str(val);
                sql += "',"
                if val == "":
                    print("", end=",")
                else:
                    print(val, end=",")
            else:
                sql += "'"
                sql += str(val)
                sql += "')"
                if val == "":
                    print("", end="")
                else:
                    print(val, end="")
        print()

        #print sql for reviewing
        print("sql="+sql);

        #run sql
        cursor.execute(sql)

        print()
        print()



    data = {
        "status":"true",
        "upload_excel":
        {
        "result": "pass",
        "full uploaded file path": newpath
        }
        }

    #commit changes to databse
    conn.commit()

    #return mysql connection to pool
    cursor.close()
    conn.close()

    await asyncio.sleep(5)

    #return json response
    return jsonify(data)

@app.route('/master_formula/listall', methods=['POST'])
def get_masterformula_listall():
    app.logger.info('/master_formula/listall')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()
    cursor.execute('select * from master_tsubakimoto_formula')
    data = cursor.fetchall()
    cursor.close()
    conn.close()
    return jsonify(data)

@app.route('/master_formula/deleteall', methods=['POST'])
def get_masterformula_deleteall():
    app.logger.info('/master_data/deleteall')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()
    cursor.execute('delete from master_tsubakimoto_formula')
    data = cursor.fetchall()
    conn.commit()
    cursor.close()
    conn.close()
    data = {
        "status":"true",
        "delete all masterdata formula":
            {
                "result": "pass"
            }
    }
    return jsonify(data)

@app.route('/master_formula/update', methods=['POST'])
def get_masterformula_update():
    app.logger.info('/master_data/update')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()

    sql = "update master_tsubakimoto_formula set "

    if request.form.get('category') is not None:
        sql += ","
        sql += "category='"
        sql += request.form.get('category')
        sql += "'"

    if request.form.get('part_no') is not None:
        sql += ","
        sql += "part_no='"
        sql += request.form.get('part_no')
        sql += "'"

    if request.form.get('previous_model_no') is not None:
        sql += ","
        sql += "previous_model_no='"
        sql += request.form.get('previous_model_no')
        sql += "'"

    if request.form.get('new_model_no') is not None:
        sql += ","
        sql += "new_model_no='"
        sql += request.form.get('new_model_no')
        sql += "'"

    if request.form.get('new_model_no') is not None:
        sql += ","
        sql += "unit='"
        sql += request.form.get('new_model_no')
        sql += "'"

    if request.form.get('manufacturer_suggested_retail_price') is not None:
        sql += ","
        sql += "manufacturer_suggested_retail_price='"
        sql += request.form.get('manufacturer_suggested_retail_price')
        sql += "'"

    if request.form.get('new_manufacturer_suggested_retail_price') is not None:
        sql += ","
        sql += "new_manufacturer_suggested_retail_price='"
        sql += request.form.get('new_manufacturer_suggested_retail_price')
        sql += "'"

    if request.form.get('conversion_to_ft') is not None:
        sql += ","
        sql += "conversion_to_ft='"
        sql += request.form.get('conversion_to_ft')
        sql += "'"

    if request.form.get('diff_for_cost') is not None:
        sql += ","
        sql += "diff_for_cost='"
        sql += request.form.get('diff_for_cost')
        sql += "'"

    if request.form.get('op_price') is not None:
        sql += ","
        sql += "op_price='"
        sql += request.form.get('op_price')
        sql += "'"

    if request.form.get('po_price_jpy_usd') is not None:
        sql += ","
        sql += "po_price_jpy_usd='"
        sql += request.form.get('po_price_jpy_usd')
        sql += "'"

    if request.form.get('po_price_currency') is not None:
        sql += ","
        sql += "po_price_currency='"
        sql += request.form.get('po_price_currency')
        sql += "'"

    if request.form.get('remark') is not None:
        sql += ","
        sql += "remark='"
        sql += request.form.get('remark')
        sql += "'"

    if request.form.get('thb_cost') is not None:
        sql += ","
        sql += "thb_cost='"
        sql += request.form.get('thb_cost')
        sql += "'"

    if request.form.get('gp') is not None:
        sql += ","
        sql += "gp='"
        sql += request.form.get('gp')
        sql += "'"

    if request.form.get('pricelist_name') is not None:
        sql += ","
        sql += "pricelist_name='"
        sql += request.form.get('pricelist_name')
        sql += "'"

    if request.form.get('multiplier') is not None:
        sql += ","
        sql += "multiplier='"
        sql += request.form.get('multiplier')
        sql += "'"

    if request.form.get('make_same_price_as_standard_price') is not None:
        sql += ","
        sql += "make_same_price_as_standard_price='"
        sql += request.form.get('make_same_price_as_standard_price')
        sql += "'"

    if request.form.get('new_make_same_price_as_standard_price') is not None:
        sql += ","
        sql += "new_make_same_price_as_standard_price='"
        sql += request.form.get('new_make_same_price_as_standard_price')
        sql += "'"

    if request.form.get('standard_price') is not None:
        sql += ","
        sql += "standard_price='"
        sql += request.form.get('standard_price')
        sql += "'"

    if request.form.get('diff') is not None:
        sql += ","
        sql += "diff='"
        sql += request.form.get('diff')
        sql += "'"

    if request.form.get('dist_pl_mul') is not None:
        sql += ","
        sql += "dist_pl_mull='"
        sql += request.form.get('dist_pl_mul')
        sql += "'"

    if request.form.get('dist_ex_rate') is not None:
        sql += ","
        sql += "dist_ex_rate='"
        sql += request.form.get('dist_ex_rate')
        sql += "'"

    if request.form.get('unit_price') is not None:
        sql += ","
        sql += "unit_price='"
        sql += request.form.get('unit_price')
        sql += "'"

    if request.form.get('new_unit_price') is not None:
        sql += ","
        sql += "new_unit_price='"
        sql += request.form.get('new_unit_price')
        sql += "'"

    if request.form.get('diff_unit_price') is not None:
        sql += ","
        sql += "diff_unit_price='"
        sql += request.form.get('diff_unit_price')
        sql += "'"

    if request.form.get('status') is not None:
        sql += ","
        sql += "status='"
        sql += request.form.get('status')
        sql += "'"

    if request.form.get('supplier_name') is not None:
        sql += ","
        sql += "supplier_name='"
        sql += request.form.get('supplier_name')
        sql += "'"

    if request.form.get('stock_reference') is not None:
        sql += ","
        sql += "stock_reference='"
        sql += request.form.get('stock_reference')
        sql += "'"

    if request.form.get('cutting_assembly') is not None:
        sql += ","
        sql += "cutting_assembly='"
        sql += request.form.get('cutting_assembly')
        sql += "'"

    if request.form.get('detail') is not None:
        sql += ","
        sql += "detail='"
        sql += request.form.get('detail')
        sql += "'"

    sql += " where Id="
    sql += request.form.get('Id')

    sql = sql.replace('update master_tsubakimoto_formula set ,','update master_tsubakimoto_formula set ')
    print('sql='+sql)

    cursor.execute(sql)
    conn.commit()

    cursor.close()
    conn.close()

    data = {
        "status":"true",
        "update masterdata":
            {
                "result": "pass"
            }
    }
    return jsonify(data)

@app.route('/master_formula/add', methods=['POST'])
def get_masterformula_add():
    app.logger.info('/master_data/update')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()

    sql = "insert into master_tsubakimoto_formula";
    sql += "(category,";
    sql += "part_no,";
    sql += "previous_model_no,";
    sql += "new_model_no,";
    sql += "unit,";
    sql += "manufacturer_suggested_retail_price,";
    sql += "new_manufacturer_suggested_retail_price,";
    sql += "conversion_to_ft,";
    sql += "diff_for_cost,";
    sql += "op_price,";
    sql += "po_price_jpy_usd,";
    sql += "po_price_currency,";
    sql += "remark,";
    sql += "thb_cost,";
    sql += "gp,";
    sql += "pricelist_name,";
    sql += "multiplier,";
    sql += "make_same_price_as_standard_price,";
    sql += "new_make_same_price_as_standard_price,";
    sql += "standard_price,";
    sql += "diff,";
    sql += "dist_pl_mull,";
    sql += "dist_ex_rate,";
    sql += "unit_price,";
    sql += "new_unit_price,";
    sql += "diff_unit_price,";
    sql += "status,";
    sql += "supplier_name,";
    sql += "stock_reference,";
    sql += "cutting_assembly,";
    sql += "detail)";
    sql += " values (";

    sql += "'"
    if request.form.get('category') is not None:
        sql += request.form.get('category')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('part_no') is not None:
        sql += request.form.get('part_no')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('previous_model_no') is not None:
        sql += request.form.get('previous_model_no')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('new_model_no') is not None:
        sql += request.form.get('new_model_no')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('unit') is not None:
        sql += request.form.get('unit')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('manufacturer_suggested_retail_price') is not None:
        sql += request.form.get('manufacturer_suggested_retail_price')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('new_manufacturer_suggested_retail_price') is not None:
        sql += request.form.get('new_manufacturer_suggested_retail_price')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('conversion_to_ft') is not None:
        sql += request.form.get('conversion_to_ft')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('diff_for_cost') is not None:
        sql += request.form.get('diff_for_cost')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('op_price') is not None:
        sql += request.form.get('op_price')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('po_price_jpy_usd') is not None:
        sql += request.form.get('po_price_jpy_usd')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('po_price_currency') is not None:
        sql += request.form.get('po_price_currency')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('remark') is not None:
        sql += request.form.get('remark')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('thb_cost') is not None:
        sql += request.form.get('thb_cost')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('gp') is not None:
        sql += request.form.get('gp')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('pricelist_name') is not None:
        sql += request.form.get('pricelist_name')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('multiplier') is not None:
        sql += request.form.get('multiplier')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('make_same_price_as_standard_price') is not None:
        sql += request.form.get('make_same_price_as_standard_price')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('new_make_same_price_as_standard_price') is not None:
        sql += request.form.get('new_make_same_price_as_standard_price')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('standard_price') is not None:
        sql += request.form.get('standard_price')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('diff') is not None:
        sql += request.form.get('diff')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('dist_pl_mull') is not None:
        sql += request.form.get('dist_pl_mull')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('dist_ex_rate') is not None:
        sql += request.form.get('dist_ex_rate')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('unit_price') is not None:
        sql += request.form.get('unit_price')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('new_unit_price') is not None:
        sql += request.form.get('new_unit_price')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('diff_unit_price') is not None:
        sql += request.form.get('diff_unit_price')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('status') is not None:
        sql += request.form.get('status')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('supplier_name') is not None:
        sql += request.form.get('supplier_name')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('stock_reference') is not None:
        sql += request.form.get('stock_reference')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('cutting_assembly') is not None:
        sql += request.form.get('cutting_assembly')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('detail') is not None:
        sql += request.form.get('detail')
    sql += "'"

    sql += ")";

    print('sql='+sql)

    cursor.execute(sql)
    conn.commit()

    cursor.close()
    conn.close()

    data = {
        "status":"true",
        "add masterdata":
            {
                "result": "pass"
            }
    }
    return jsonify(data)

@app.route('/master_formula/delete', methods=['POST'])
def get_masterformula_delete():
    app.logger.info('/master_formula/delete')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()

    sql = "delete from master_tsubakimoto_formula"
    sql += " where Id="
    sql += request.form.get('Id')

    print('sql='+sql)

    cursor.execute(sql)
    conn.commit()

    cursor.close()
    conn.close()

    data = {
        "status":"true",
        "delete masterformula with id":
            {
                "result": "pass"
            }
    }
    return jsonify(data)

@app.route('/exchange_rate/upload', methods=['POST'])
async def get_exchangerate_upload():
    app.logger.info('/master_formula/upload')

    #request mysql connection from pool
    conn = connection_pool.get_connection()
    cursor = conn.cursor()

    # upload file
    file = request.files['file']
    fullfilename = file.filename
    onlyfilename = fullfilename.split('.')[0];
    onlyfilename = onlyfilename.replace(' ','_')
    onlyfilename = onlyfilename.replace('-','_')
    onlyfileext = fullfilename.split('.')[1];
    print(request.files);
    newpath = "uploaded_files/" + onlyfilename  + "_" + datetime.now(pytz.timezone('Asia/Bangkok')).strftime('%Y_%m_%d_%H_%M_%S') + "." + onlyfileext;
    app.logger.info("uploaded new file path : "+newpath)
    file.save(newpath)

    # parse file
    wb = openpyxl.load_workbook(newpath,data_only=True)
    ws = wb.active
    print('Total number of rows: '+str(ws.max_row)+'. And total number of columns: '+str(ws.max_column))

    usd_br = ws.cell(2,2).value
    print('usd_br='+str(usd_br))
    eur_br = ws.cell(2,3).value
    print('eur_br='+str(eur_br))
    jpy_br = ws.cell(2,4).value
    print('jpy_br='+str(jpy_br))

    usd_cr = ws.cell(3,2).value
    print('usd_cr='+str(usd_cr))
    eur_cr = ws.cell(3,3).value
    print('eur_cr='+str(eur_cr))
    jpy_cr = ws.cell(3,4).value
    print('jpy_cr='+str(jpy_cr))

    usd_pr = ws.cell(4,2).value
    print('usd_pr='+str(usd_pr))
    eur_pr = ws.cell(4,3).value
    print('eur_pr='+str(eur_pr))
    jpy_pr = ws.cell(4,4).value
    print('jpy_pr='+str(jpy_pr))

    usd_qr = ws.cell(5,2).value
    print('usd_qr='+str(usd_qr))
    eur_qr = ws.cell(5,3).value
    print('eur_qr='+str(eur_qr))
    jpy_qr = ws.cell(5,4).value
    print('jpy_qr='+str(jpy_qr))

    remark = ws.cell(6,2).value

    sql="insert into exchange_rate(usd_br,usd_cr,usd_pr,usd_qr,eur_br,eur_cr,eur_qr,eur_pr,jpy_br,jpy_cr,jpy_pr,jpy_qr,rate_remark,rate_file_name,rate_path)"
    sql += " values ("
    sql += str(usd_br)
    sql += ","
    sql += str(usd_cr)
    sql += ","
    sql += str(usd_pr)
    sql += ","
    sql += str(usd_qr)
    sql += ","
    sql += str(eur_br)
    sql += ","
    sql += str(eur_cr)
    sql += ","
    sql += str(eur_pr)
    sql += ","
    sql += str(eur_qr)
    sql += ","
    sql += str(jpy_br)
    sql += ","
    sql += str(jpy_cr)
    sql += ","
    sql += str(jpy_pr)
    sql += ","
    sql += str(jpy_qr)
    sql += ",'"
    sql += remark
    sql += "','"
    sql += onlyfilename
    sql += "','"
    sql += newpath
    sql += "')"

    #print sql for reviewing
    print("sql="+sql);

    #run sql
    cursor.execute(sql)

    print()
    print()


    data = {
        "status":"true",
        "upload_excel":
        {
        "result": "pass",
        "full uploaded file path": newpath
        }
        }

    #commit changes to databse
    conn.commit()

    #return mysql connection to pool
    cursor.close()
    conn.close()

    await asyncio.sleep(5)

    #return json response
    return jsonify(data)

@app.route('/exchange_rate/listall', methods=['POST'])
def get_exchangerate_listall():
    app.logger.info('/exchange_rate/listall')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()
    cursor.execute('select * from exchange_rate')
    data = cursor.fetchall()
    cursor.close()
    conn.close()
    return jsonify(data)

@app.route('/exchange_rate/deleteall', methods=['POST'])
def get_exchangerate_deleteall():
    app.logger.info('/exchange_rate/deleteall')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()
    cursor.execute('delete from exchange_rate')
    data = cursor.fetchall()
    cursor.close()
    conn.close()
    data = {
        "status":"true",
        "delete all exchange_rate":
            {
                "result": "pass"
            }
    }
    return jsonify(data)

@app.route('/exchange_rate/update', methods=['POST'])
def get_exchangerate_update():
    app.logger.info('/exchange_rate/update')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()

    sql = "update exchange_rate set "

    if request.form.get('usd_br') is not None:
        sql += ","
        sql += "usd_br="
        sql += request.form.get('usd_br')

    if request.form.get('usd_cr') is not None:
        sql += ","
        sql += "usd_cr="
        sql += request.form.get('usd_cr')

    if request.form.get('usd_qr') is not None:
        sql += ","
        sql += "usd_qr="
        sql += request.form.get('usd_qr')

    if request.form.get('usd_pr') is not None:
        sql += ","
        sql += "usd_pr="
        sql += request.form.get('usd_pr')

    if request.form.get('eur_br') is not None:
        sql += ","
        sql += "eur_br="
        sql += request.form.get('eur_br')

    if request.form.get('eur_cr') is not None:
        sql += ","
        sql += "eur_cr="
        sql += request.form.get('eur_cr')

    if request.form.get('eur_qr') is not None:
        sql += ","
        sql += "eur_qr="
        sql += request.form.get('eur_qr')

    if request.form.get('eur_pr') is not None:
        sql += ","
        sql += "eur_pr="
        sql += request.form.get('eur_pr')

    if request.form.get('jpy_br') is not None:
        sql += ","
        sql += "jpy_br="
        sql += request.form.get('jpy_br')

    if request.form.get('jpy_cr') is not None:
        sql += ","
        sql += "jpy_cr="
        sql += request.form.get('jpy_cr')

    if request.form.get('jpy_qr') is not None:
        sql += ","
        sql += "jpy_qr="
        sql += request.form.get('jpy_qr')

    if request.form.get('jpy_pr') is not None:
        sql += ","
        sql += "jpy_pr="
        sql += request.form.get('jpy_pr')

    if request.form.get('rate_remark') is not None:
        sql += ","
        sql += "rate_remark='"
        sql += request.form.get('rate_remark')
        sql += "'"

    if request.form.get('file_name') is not None:
        sql += ","
        sql += "file_name='"
        sql += request.form.get('file_name')
        sql += "'"

    if request.form.get('rate_path') is not None:
        sql += ","
        sql += "rate_path='"
        sql += request.form.get('rate_path')
        sql += "'"

    sql += " where rate_id="
    sql += request.form.get('rate_id')

    sql = sql.replace("update exchange_rate set ,", "update exchange_rate set ");

    print('sql='+sql)

    cursor.execute(sql)
    conn.commit()

    cursor.close()
    conn.close()

    data = {
        "status":"true",
        "update exchange_rate":
            {
                "result": "pass"
            }
    }
    return jsonify(data)

@app.route('/exchange_rate/delete', methods=['POST'])
def get_exchangerate_delete():
    app.logger.info('/exchange_rate/delete')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()

    sql = "delete from exchange_rate"
    sql += " where rate_id="
    sql += request.form.get('rate_id')

    print('sql='+sql)

    cursor.execute(sql)
    conn.commit()

    cursor.close()
    conn.close()

    data = {
        "status":"true",
        "delete exchange_rate by id":
            {
                "result": "pass"
            }
    }
    return jsonify(data)

@app.route('/exchange_rate/add', methods=['POST'])
def get_exchangerate_add():
    app.logger.info('/exchange_rate/add')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()

    sql = "insert into exchange_rate(usd_br,usd_cr,usd_pr,usd_qr,eur_br,eur_cr,eur_qr,eur_pr,jpy_br,jpy_cr,jpy_pr,jpy_qr,rate_remark,rate_file_name,rate_path)";
    sql += " values ("

    if request.form.get('usd_br') is not None:
       sql += request.form.get('usd_br')
    else:
       sql += "null"

    sql += ","

    if request.form.get('usd_cr') is not None:
        sql += request.form.get('usd_cr')
    else:
        sql += "null"

    sql += ","

    if request.form.get('usd_pr') is not None:
        sql += request.form.get('usd_pr')
    else:
        sql += "null"

    sql += ","

    if request.form.get('usd_qr') is not None:
        sql += request.form.get('usd_qr')
    else:
        sql += "null"

    sql += ","

    if request.form.get('eur_br') is not None:
        sql += request.form.get('eur_br')
    else:
        sql += "null"

    sql += ","

    if request.form.get('eur_cr') is not None:
        sql += request.form.get('eur_cr')
    else:
        sql += "null"

    sql += ","

    if request.form.get('eur_qr') is not None:
        sql += request.form.get('eur_qr')
    else:
        sql += "null"

    sql += ","

    if request.form.get('eur_pr') is not None:
        sql += request.form.get('eur_pr')
    else:
        sql += "null"

    sql += ","

    if request.form.get('jpy_br') is not None:
        sql += request.form.get('jpy_br')
    else:
        sql += "null"

    sql += ","

    if request.form.get('jpy_cr') is not None:
        sql += request.form.get('jpy_cr')
    else:
        sql += "null"

    sql += ","

    if request.form.get('jpy_pr') is not None:
        sql += request.form.get('jpy_pr')
    else:
        sql += "null"

    sql += ","

    if request.form.get('jpy_qr') is not None:
        sql += request.form.get('jpy_qr')
    else:
        sql += "null"

    sql += ","

    sql += "'"
    if request.form.get('rate_remark') is not None:
        sql += request.form.get('rate_remark');
    sql += "'"

    sql += ","

    sql += "'"
    if request.form.get('rate_file_name') is not None:
        sql += request.form.get('rate_file_name');
    sql += "'"

    sql += ","

    sql += "'"
    if request.form.get('rate_path') is not None:
        sql += request.form.get('rate_path');
    sql += "'"

    sql += ")"

    print('sql='+sql)

    cursor.execute(sql)
    conn.commit()

    cursor.close()
    conn.close()

    data = {
        "status":"true",
        "add exchange_rate":
            {
                "result": "pass"
            }
    }
    return jsonify(data)

@app.route('/master_history/listall', methods=['POST'])
def get_masterhistory_listall():
    app.logger.info('/master_history/listall')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()
    cursor.execute('select * from master_pricelist_history')
    data = cursor.fetchall()
    cursor.close()
    conn.close()
    return jsonify(data)

@app.route('/master_history/add', methods=['POST'])
def get_masterhistory_add():
    app.logger.info('/master_history/add')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()

    sql = "insert into master_pricelist_history(master_file_name,file_path)"
    sql += " values ("

    if request.form.get('master_file_name') is not None:
        sql += "'" + request.form.get('master_file_name') + "'"
    else:
        sql += "null"

    sql += ","

    if request.form.get('file_path') is not None:
        sql += "'" + request.form.get('file_path') + "'";
    else:
        sql += "null";

    sql += ")"

    print('sql='+sql)

    cursor.execute(sql)
    conn.commit()

    cursor.close()
    conn.close()

    data = {
        "status":"true",
        "add master_history":
            {
                "result": "pass"
            }
    }
    return jsonify(data)

@app.route('/master_history/update', methods=['POST'])
def get_masterhistory_update():
    app.logger.info('/master_history/update')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()

    sql = "update master_pricelist_history set ";

    if request.form.get('master_file_name') is not None:
        sql += ","
        sql += "master_file_name='"
        sql += request.form.get('master_file_name')
        sql += "'"


    if request.form.get('file_path') is not None:
        sql += ",";
        sql += "file_path='";
        sql += request.form.get('file_path')
        sql += "'";

    sql += " where master_file_id="
    sql += request.form.get('master_file_id')

    sql = sql.replace("update master_pricelist_history set ,", "update master_pricelist_history set ")

    print('sql='+sql)

    cursor.execute(sql)
    conn.commit()

    cursor.close()
    conn.close()

    data = {
        "status":"true",
        "update master_history":
            {
                "result": "pass"
            }
    }
    return jsonify(data)

@app.route('/master_history/delete', methods=['POST'])
def get_masterhistory_delete():
    app.logger.info('/master_history/delete')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()

    sql = "delete from master_pricelist_history";
    sql += " where master_file_id="
    sql += request.form.get('master_file_id')

    print('sql='+sql)

    cursor.execute(sql)
    conn.commit()

    cursor.close()
    conn.close()

    data = {
        "status":"true",
        "delete master_history":
            {
                "result": "pass"
            }
    }
    return jsonify(data)

@app.route('/user/listall', methods=['POST'])
def get_user_listall():
    app.logger.info('/user/listall')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()
    cursor.execute('select * from user')
    data = cursor.fetchall()
    cursor.close()
    conn.close()
    return jsonify(data)

@app.route('/user/delete', methods=['POST'])
def get_user_delete():
    app.logger.info('/user/delete')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()
    sql = "delete from user where user_id="+request.form.get('id')
    cursor.execute(sql)
    conn.commit()
    cursor.close()
    conn.close()
    data = {
        "status":"true",
        "delete user":
            {
                "result": "pass"
            }
    }
    return jsonify(data)

@app.route('/user/add', methods=['POST'])
def get_user_add():
    app.logger.info('/user/add')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()
    sql = "insert into user(email,password,access_type,name_surname,company_name) values ('" + request.form.get('usr') + "','" + request.form.get('pwd') + "','" + request.form.get('access') + "','" + request.form.get('name') + "','" + request.form.get('company') + "')";
    cursor.execute(sql)
    conn.commit()
    cursor.close()
    conn.close()
    data = {
        "status":"true",
        "add user":
            {
                "result": "pass"
            }
    }
    return jsonify(data)

@app.route('/user/update', methods=['POST'])
def get_user_update():
    app.logger.info('/user/update')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()
    sql = "update user set email='"+request.form.get('usr')+"',password='"+request.form.get('pwd')+"',access_type='"+request.form.get('access')+"',name_surname='"+request.form.get('name')+"',company_name='"+request.form.get('company')+"' where user_id="+request.form.get('id')
    cursor.execute(sql)
    conn.commit()
    cursor.close()
    conn.close()
    data = {
        "status":"true",
        "update user":
            {
                "result": "pass"
            }
    }
    return jsonify(data)

@app.route('/master_tmp/upload', methods=['POST'])
async def get_mastertmp_upload():
    app.logger.info('/master_tmp/upload')

    #request mysql connection from pool
    conn = connection_pool.get_connection()
    cursor = conn.cursor()

    # upload file
    file = request.files['file']
    fullfilename = file.filename
    onlyfilename = fullfilename.split('.')[0];
    onlyfilename = onlyfilename.replace(' ','_')
    onlyfilename = onlyfilename.replace('-','_')
    onlyfileext = fullfilename.split('.')[1];
    print(request.files);
    newpath = "uploaded_files/" + onlyfilename  + "_" + datetime.now(pytz.timezone('Asia/Bangkok')).strftime('%Y_%m_%d_%H_%M_%S') + "." + onlyfileext;
    app.logger.info("uploaded new file path : "+newpath)
    file.save(newpath)

    # requests.post('http://deploy-aws.com:5000/master_history/add', data={"master_file_name":onlyfilename  + "_" + datetime.now(pytz.timezone('Asia/Bangkok')).strftime('%Y_%m_%d_%H_%M_%S') + "." + onlyfileext,"file_path":newpath})

    # parse file
    wb = openpyxl.load_workbook(newpath,data_only=True)
    ws = wb.active
    print('Total number of rows: '+str(ws.max_row)+'. And total number of columns: '+str(ws.max_column))
    for row in range(5, ws.max_row+1):
        sql="insert into master_tsubakimoto_tmp(category,part_no,previous_model_no,new_model_no,unit,manufacturer_suggested_retail_price,new_manufacturer_suggested_retail_price,conversion_to_ft,diff_for_cost,op_price,po_price_jpy_usd,po_price_currency,remark,thb_cost,gp,pricelist_name,multiplier,make_same_price_as_standard_price,new_make_same_price_as_standard_price,standard_price,diff,dist_pl_mull,dist_ex_rate,unit_price,new_unit_price,diff_unit_price,status,supplier_name,stock_reference,cutting_assembly,detail)";
        sql += " values (";
        for column in range(1, ws.max_column+1):
            val = ws.cell(row,column).value
            if val is str:
                val = val.replace('\n','')
                val = val.replace('\r','')
                val = val.replace('\t','')
            elif val is None or val == '#VALUE!':
                val = "";
            if column < ws.max_column:
                sql += "'"
                sql += str(val);
                sql += "',"
                if val == "":
                    print("", end=",")
                else:
                    print(val, end=",")
            else:
                sql += "'"
                sql += str(val)
                sql += "')"
                if val == "":
                    print("", end="")
                else:
                    print(val, end="")
        print()

        #print sql for reviewing
        print("sql="+sql);

        #run sql
        cursor.execute(sql)

        print()
        print()



    data = {
        "status":"true",
        "upload_master_tmp":
            {
                "result": "pass",
                "full uploaded file path": newpath
            }
    }

    #commit changes to databse
    conn.commit()

    #return mysql connection to pool
    cursor.close()
    conn.close()

    await asyncio.sleep(5)

    #return json response
    return jsonify(data)

@app.route('/master_tmp/listall', methods=['POST'])
def get_mastertmp_listall():
    app.logger.info('/master_tmp/listall')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()
    cursor.execute('select * from master_tsubakimoto_tmp')
    data = cursor.fetchall()
    cursor.close()
    conn.close()
    return jsonify(data)

@app.route('/master_tmp/deleteall', methods=['POST'])
def get_mastertmp_deleteall():
    app.logger.info('/master_tmp/deleteall')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()
    cursor.execute('delete from master_tsubakimoto_tmp')
    conn.commit()
    cursor.close()
    conn.close()

    data = {
        "status":"true",
        "delete_master_tmp":
            {
                "result": "pass"
            }
    }

    return jsonify(data)

@app.route('/company/listall', methods=['POST'])
def get_company_listall():
    app.logger.info('/company/listall')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()
    cursor.execute('select * from company')
    data = cursor.fetchall()
    cursor.close()
    conn.close()
    return jsonify(data)

@app.route('/company/update', methods=['POST'])
def get_company_update():
    app.logger.info('/company/update')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()


    sql = "update company set ";

    if request.form.get('company_name') is not None:
        sql += ","
        sql += "company_name='"
        sql += request.form.get('company_name')
        sql += "'"

    if request.form.get('company_info') is not None:
        sql += ","
        sql += "company_info='"
        sql += request.form.get('company_info')
        sql += "'"

    if request.form.get('company_phone') is not None:
        sql += ","
        sql += "company_phone='"
        sql += request.form.get('company_phone')
        sql += "'"

    if request.form.get('company_fax') is not None:
        sql += ","
        sql += "company_fax='"
        sql += request.form.get('company_fax')
        sql += "'"

    if request.form.get('company_email') is not None:
        sql += ","
        sql += "company_email='"
        sql += request.form.get('company_email')
        sql += "'"

    if request.form.get('company_pic') is not None:
        sql += ","
        sql += "company_pic='"
        sql += request.form.get('company_pic')
        sql += "'"
    sql += " where company_id="
    sql += request.form.get('company_id')
    sql = sql.replace("update company set ,", "update company set ")

    cursor.execute(sql)
    conn.commit()
    cursor.close()
    conn.close()
    data = {
        "status":"true",
        "update company":
            {
                "result": "pass"
            }
    }
    return jsonify(data)

@app.route('/company/add', methods=['POST'])
def get_company_add():
    app.logger.info('/company/add')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()

    sql = "insert into company";
    sql += "(company_name,";
    sql += "company_info,";
    sql += "company_phone,";
    sql += "company_fax,";
    sql += "company_email,";
    sql += "company_pic)";
    sql += " values (";

    sql += "'"
    if request.form.get('company_name') is not None:
     sql += request.form.get('company_name')
    sql += "'"

    sql += ","

    sql += "'"
    if request.form.get('company_info') is not None:
        sql += request.form.get('company_info')
    sql += "'"

    sql += ","

    sql += "'"
    if request.form.get('company_phone') is not None:
        sql += request.form.get('company_phone')
    sql += "'"

    sql += ","

    sql += "'"
    if request.form.get('company_fax') is not None:
        sql += request.form.get('company_fax')
    sql += "'"

    sql += ","

    sql += "'"
    if request.form.get('company_email') is not None:
        sql += request.form.get('company_email')
    sql += "'"

    sql += ","

    sql += "'"
    if request.form.get('company_pic') is not None:
        sql += request.form.get('company_pic')
    sql += "'"

    sql += ")"

    cursor.execute(sql)
    conn.commit()
    cursor.close()
    conn.close()
    data = {
        "status":"true",
        "add company":
            {
                "result": "pass"
            }
    }
    return jsonify(data)

@app.route('/company/delete', methods=['POST'])
def get_company_delete():
    app.logger.info('/company/update')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()

    sql = "delete from company";
    sql += " where company_id="
    sql += request.form.get('company_id')

    cursor.execute(sql)
    conn.commit()
    cursor.close()
    conn.close()
    data = {
        "status":"true",
        "delete company":
            {
                "result": "pass"
            }
    }
    return jsonify(data)

@app.route('/distributor/listall', methods=['POST'])
def get_distributor_listall():
    app.logger.info('/distributor/listall')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()
    cursor.execute('select * from distributor_product_matching')
    data = cursor.fetchall()
    cursor.close()
    conn.close()
    return jsonify(data)

@app.route('/distributor/update', methods=['POST'])
def get_distributor_update():
    app.logger.info('/distributor/update')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()

    sql = "update distributor_product_matching set ";

    if request.form.get('master_price_list') is not None:
        sql += ","
        sql += "master_price_list='"
        sql += request.form.get('master_price_list')
        sql += "'"

    if request.form.get('master_pricelist_showing_name') is not None:
        sql += ","
        sql += "master_pricelist_showing_name='"
        sql += request.form.get('master_pricelist_showing_name')
        sql += "'"

    if request.form.get('company_id') is not None:
        sql += ","
        sql += "company_id="
        sql += request.form.get('company_id')

    sql += " where mc_id="
    sql += request.form.get('mc_id')
    sql = sql.replace("update distributor_product_matching set ,", "update distributor_product_matching set ")

    cursor.execute(sql)
    conn.commit()
    cursor.close()
    conn.close()
    data = {
        "status":"true",
        "update distributor":
            {
                "result": "pass"
            }
    }
    return jsonify(data)

@app.route('/distributor/add', methods=['POST'])
def get_distributor_add():
    app.logger.info('/distributor/add')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()

    sql = "insert into distributor_product_matching";
    sql += "(company_id,"
    sql += "master_price_list,"
    sql += "master_pricelist_showing_name)"
    sql += " values (";

    if request.form.get('company_id') is not None:
        sql += request.form.get('company_id')
    else:
        sql += "null"

    sql += ","

    sql += "'"
    if request.form.get('master_price_list') is not None:
        sql += request.form.get('master_price_list')
    sql += "'"

    sql += ","

    sql += "'"
    if request.form.get('master_pricelist_showing_name') is not None:
        sql += request.form.get('master_pricelist_showing_name')
    sql += "'"

    sql += ")"

    cursor.execute(sql)
    conn.commit()
    cursor.close()
    conn.close()
    data = {
        "status":"true",
        "add distributor":
            {
                "result": "pass"
            }
    }
    return jsonify(data)

@app.route('/distributor/delete', methods=['POST'])
def get_distributor_delete():
    app.logger.info('/distributor/delete')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()

    sql = "delete from distributor_product_matching";
    sql += " where mc_id="
    sql += request.form.get('mc_id')

    cursor.execute(sql)
    conn.commit()
    cursor.close()
    conn.close()
    data = {
        "status":"true",
        "delete distributor":
            {
                "result": "pass"
            }
    }
    return jsonify(data)

@app.route('/supplier/listall', methods=['POST'])
def get_supplier_listall():
    app.logger.info('/supplier/listall')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()
    cursor.execute('select * from supplier_matching')
    data = cursor.fetchall()
    cursor.close()
    conn.close()
    return jsonify(data)

@app.route('/supplier/add', methods=['POST'])
def get_supplier_add():
    app.logger.info('/supplier/add')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()

    sql = "insert into supplier_matching"
    sql += "(sup_name,"
    sql += "master_pricelist_name,sup_short_name)"
    sql += " values ("

    sql += "'"
    if request.form.get('sup_name') is not None:
        sql += request.form.get('sup_name')
    sql += "'"

    sql += ","

    sql += "'"
    if request.form.get('master_pricelist_name') is not None:
        sql += request.form.get('master_pricelist_name')
    sql += "'"

    sql += ","

    sql += "'"
    if request.form.get('sup_short_name') is not None:
        sql += request.form.get('sup_short_name')
    sql += "'"

    sql += ")"

    cursor.execute(sql)
    conn.commit()
    cursor.close()
    conn.close()
    data = {
        "status":"true",
        "add supplier":
            {
                "result": "pass"
            }
    }
    return jsonify(data)

@app.route('/supplier/delete', methods=['POST'])
def get_supplier_delete():
    app.logger.info('/supplier/delete')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()

    sql = "delete from supplier_matching";
    sql += " where sup_id="
    sql += request.form.get('sup_id')

    cursor.execute(sql)
    conn.commit()
    cursor.close()
    conn.close()
    data = {
        "status":"true",
        "delete supplier":
            {
                "result": "pass"
            }
    }
    return jsonify(data)

@app.route('/exchange_rate_history/listall', methods=['POST'])
def get_exchange_rate_history_listall():
    app.logger.info('/exchange_rate_history/listall')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()
    cursor.execute('select * from exchange_rate_history')
    data = cursor.fetchall()
    cursor.close()
    conn.close()
    return jsonify(data)

@app.route('/exchange_rate_history/delete', methods=['POST'])
def get_exchange_rate_history_delete():
    app.logger.info('/exchange_rate_history/delete')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()

    sql = "delete from exchange_rate_history";
    sql += " where rate_doc_id="
    sql += request.form.get('rate_doc_id')

    cursor.execute(sql)
    conn.commit()
    cursor.close()
    conn.close()
    data = {
        "status":"true",
        "delete exchange_rate_history":
            {
                "result": "pass"
            }
    }
    return jsonify(data)

@app.route('/exchange_rate_history/upload', methods=['POST'])
def get_exchange_rate_history_upload():
    app.logger.info('/exchange_rate_history/upload')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()

    sql = "insert into exchange_rate_history(rate_doc_name,rate_doc_path) values (";

    sql += "'"
    if request.form.get('rate_doc_name') is not None:
     sql += request.form.get('rate_doc_name')
    sql += "'"

    sql += ","

    sql += "'"
    if request.form.get('rate_doc_path') is not None:
        sql += request.form.get('rate_doc_path')
    sql += "'"

    sql += ")"

    cursor.execute(sql)
    conn.commit()
    cursor.close()
    conn.close()
    data = {
        "status":"true",
        "upload exchange_rate_history":
            {
                "result": "pass"
            }
    }
    return jsonify(data)

@app.route('/exchange_rate_history/download', methods=['POST'])
def get_exchange_rate_history_download():
    app.logger.info('/exchange_rate_history/download')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()

    sql = "select * from exchange_rate_history where "
    sql += "rate_doc_id=";
    sql += request.form.get('rate_doc_id');
    cursor.execute(sql)
    results = cursor.fetchall()
    cursor.close()
    conn.close()
    data = {
        "status":"true",
        "show download link":
            {
                "document_name": results[0].rate_doc_name,
                "document_path": results[0].rate_doc_path,
                "document_download_link": "http://deploy-aws.com:3006/downloadfiletocomputer?fileurl="+results[0].rate_doc_name
            }
    }
    return jsonify(data)

@app.route('/cost_history/listall', methods=['POST'])
def get_costhistory_listall():
    app.logger.info('/cost_history/listall')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()
    cursor.execute('select * from cost_file_history')
    data = cursor.fetchall()
    cursor.close()
    conn.close()
    return jsonify(data)

@app.route('/cost_history/add', methods=['POST'])
def get_coshistory_add():
    app.logger.info('/cost_history/add')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()

    sql = "insert into cost_file_history";
    sql += "(cost_file_name,"
    sql += "im_path)"
    sql += " values ("

    sql += "'"
    if request.form.get('cost_file_name') is not None:
        sql += request.form.get('cost_file_name')
    sql += "'"

    sql += ","

    sql += "'"
    if request.form.get('im_path') is not None:
        sql += request.form.get('im_path')
    sql += "'"

    sql += ")"

    cursor.execute(sql)
    conn.commit()
    cursor.close()
    conn.close()
    data = {
        "status":"true",
        "add cost_history":
            {
                "result": "pass"
            }
    }
    return jsonify(data)

@app.route('/cost_history/delete', methods=['POST'])
def get_costhistory_delete():
    app.logger.info('/cost_history/delete')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()

    sql = "delete from cost_file_history";
    sql += " where im_cost_id="
    sql += request.form.get('im_cost_id')

    cursor.execute(sql)
    conn.commit()
    cursor.close()
    conn.close()
    data = {
        "status":"true",
        "delete cost_history":
            {
                "result": "pass"
            }
    }
    return jsonify(data)

@app.route('/update_master/listall', methods=['POST'])
def get_updatemaster_listall():
    app.logger.info('/update_master/listall')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()
    cursor.execute('select * from updating_master_price_list')
    data = cursor.fetchall()
    cursor.close()
    conn.close()
    return jsonify(data)

@app.route('/update_master/add', methods=['POST'])
def get_updatemaster_add():
    app.logger.info('/update_master/add')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()

    sql = "insert into updating_master_price_list";
    sql += "(category,";
    sql += "part_no,";
    sql += "previous_model_no,";
    sql += "new_model_no,";
    sql += "unit,";
    sql += "manufacturer_suggested_retail_price,";
    sql += "new_manufacturer_suggested_retail_price,";
    sql += "conversion_to_ft,";
    sql += "diff_for_cost,";
    sql += "op_price,";
    sql += "po_price_jpy_usd,";
    sql += "po_price_currency,";
    sql += "remark,";
    sql += "thb_cost,";
    sql += "gp,";
    sql += "pricelist_name,";
    sql += "multiplier,";
    sql += "make_same_price_as_standard_price,";
    sql += "new_make_same_price_as_standard_price,";
    sql += "standard_price,";
    sql += "diff,";
    sql += "dist_pl_mull,";
    sql += "dist_ex_rate,";
    sql += "unit_price,";
    sql += "new_unit_price,";
    sql += "diff_unit_price,";
    sql += "status,";
    sql += "supplier_name,";
    sql += "stock_reference,";
    sql += "cutting_assembly,";
    sql += "detail,Id)";
    sql += " values (";

    sql += "'"
    if request.form.get('category') is not None:
        sql += request.form.get('category')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('part_no') is not None:
        sql += request.form.get('part_no')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('previous_model_no') is not None:
        sql += request.form.get('previous_model_no')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('new_model_no') is not None:
        sql += request.form.get('new_model_no')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('unit') is not None:
        sql += request.form.get('unit')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('manufacturer_suggested_retail_price') is not None:
        sql += request.form.get('manufacturer_suggested_retail_price')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('new_manufacturer_suggested_retail_price') is not None:
        sql += request.form.get('new_manufacturer_suggested_retail_price')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('conversion_to_ft') is not None:
        sql += request.form.get('conversion_to_ft')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('diff_for_cost') is not None:
        sql += request.form.get('diff_for_cost')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('op_price') is not None:
        sql += request.form.get('op_price')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('po_price_jpy_usd') is not None:
        sql += request.form.get('po_price_jpy_usd')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('po_price_currency') is not None:
        sql += request.form.get('po_price_currency')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('remark') is not None:
        sql += request.form.get('remark')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('thb_cost') is not None:
        sql += request.form.get('thb_cost')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('gp') is not None:
        sql += request.form.get('gp')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('pricelist_name') is not None:
        sql += request.form.get('pricelist_name')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('multiplier') is not None:
        sql += request.form.get('multiplier')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('make_same_price_as_standard_price') is not None:
        sql += request.form.get('make_same_price_as_standard_price')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('new_make_same_price_as_standard_price') is not None:
        sql += request.form.get('new_make_same_price_as_standard_price')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('standard_price') is not None:
        sql += request.form.get('standard_price')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('diff') is not None:
        sql += request.form.get('diff')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('dist_pl_mull') is not None:
        sql += request.form.get('dist_pl_mull')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('dist_ex_rate') is not None:
        sql += request.form.get('dist_ex_rate')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('unit_price') is not None:
        sql += request.form.get('unit_price')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('new_unit_price') is not None:
        sql += request.form.get('new_unit_price')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('diff_unit_price') is not None:
        sql += request.form.get('diff_unit_price')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('status') is not None:
        sql += request.form.get('status')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('supplier_name') is not None:
        sql += request.form.get('supplier_name')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('stock_reference') is not None:
        sql += request.form.get('stock_reference')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('cutting_assembly') is not None:
        sql += request.form.get('cutting_assembly')
    sql += "'"

    sql += ','

    sql += "'"
    if request.form.get('detail') is not None:
        sql += request.form.get('detail')
    sql += "'"

    sql += ','
    if request.form.get('Id') is not None:
        sql += request.form.get('Id')


    sql += ")";

    print('sql='+sql)

    cursor.execute(sql)
    conn.commit()

    cursor.close()
    conn.close()

    data = {
        "status":"true",
        "add masterdata":
            {
                "result": "pass"
            }
    }
    return jsonify(data)

app.route('/update_master/delete', methods=['POST'])
def get_updatemaster_delete():
    app.logger.info('/update_master/delete')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()

    sql = "delete from updating_master_price_list";
    sql += " where Id="
    sql += request.form.get('Id')

    cursor.execute(sql)
    conn.commit()
    cursor.close()
    conn.close()
    data = {
        "status":"true",
        "delete updating_master_price_list by id":
            {
                "result": "pass"
            }
    }
    return jsonify(data)

@app.route('/cost/listall', methods=['POST'])
def get_cost_listall():
    app.logger.info('/cost/listall')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()
    cursor.execute('select * from cost')
    data = cursor.fetchall()
    cursor.close()
    conn.close()
    return jsonify(data)

@app.route('/cost/deleteall', methods=['POST'])
def get_cost_deleteall():
    app.logger.info('/cost/deleteall')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()
    cursor.execute('delete from cost')
    # data = cursor.fetchall()
    conn.commit()
    cursor.close()
    conn.close()
    data = {
        "status":"true",
        "delete all cost":
            {
                "result": "pass"
            }
    }
    return jsonify(data)

@app.route('/cost/upload', methods=['POST'])
async def get_cost_upload():
    app.logger.info('/cost/upload')

    #request mysql connection from pool
    conn = connection_pool.get_connection()
    cursor = conn.cursor()

    # upload file
    file = request.files['file']
    fullfilename = file.filename
    onlyfilename = fullfilename.split('.')[0];
    onlyfilename = onlyfilename.replace(' ','_')
    onlyfilename = onlyfilename.replace('-','_')
    onlyfileext = fullfilename.split('.')[1];
    print(request.files);
    newpath = "uploaded_files/" + onlyfilename  + "_" + datetime.now(pytz.timezone('Asia/Bangkok')).strftime('%Y_%m_%d_%H_%M_%S') + "." + onlyfileext;
    app.logger.info("uploaded new file path : "+newpath)
    file.save(newpath)

    # parse file
    wb = openpyxl.load_workbook(newpath,data_only=True)
    ws = wb.active
    print('Total number of rows: '+str(ws.max_row)+'. And total number of columns: '+str(ws.max_column))
    for row in range(3, ws.max_row+1):
        sql="insert into cost(category,part_no,model_no,unit,manufacturer_suggested_retail_price,sub_price_list)";
        sql += " values (";
        # for column in range(1, ws.max_column+1):
        for column in range(1, 7):
            val = ws.cell(row,column).value
            if val is str:
                val = val.replace('\n','')
                val = val.replace('\r','')
                val = val.replace('\t','')
            elif val is None or val == '#VALUE!':
                val = "";
            if column < 6:
                sql += "'"
                sql += str(val);
                sql += "',"
                if val == "":
                    print("", end=",")
                else:
                    print(val, end=",")
            else:
                sql += "'"
                sql += str(val)
                sql += "')"
                if val == "":
                    print("", end="")
                else:
                    print(val, end="")
        print()

        #print sql for reviewing
        print("sql="+sql);

        #run sql
        cursor.execute(sql)

        print()
        print()



    data = {
        "status":"true",
        "upload_cost":
            {
                "result": "pass",
                "full uploaded file path": newpath
            }
    }

    #commit changes to databse
    conn.commit()

    #return mysql connection to pool
    cursor.close()
    conn.close()

    await asyncio.sleep(5)

    #return json response
    return jsonify(data)

@app.route('/quotation_file_record/listall', methods=['POST'])
def get_quotationfilerecord_listall():
    app.logger.info('/quotation_file_record/listall')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()
    cursor.execute('select * from quotation_file_record')
    data = cursor.fetchall()
    cursor.close()
    conn.close()
    return jsonify(data)

app.route('/quotation_file_record/delete', methods=['POST'])
def get_quotationfilerecord_delete():
    app.logger.info('/quotation_file_record/delete')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()

    sql = "delete from quotation_file_record where ";
    sql += "quot_file_id=";
    sql += request.form.get('quot_file_id');

    cursor.execute(sql)
    conn.commit()
    cursor.close()
    conn.close()
    data = {
        "status":"true",
        "delete quotationfilerecord by id":
            {
                "result": "pass"
            }
    }
    return jsonify(data)

@app.route('/quotation_file_record/upload', methods=['POST'])
async def get_quotationfilerecord_upload():
    app.logger.info('/quotation_file_record/upload')

    # upload file
    file = request.files['file']
    fullfilename = file.filename
    onlyfilename = fullfilename.split('.')[0];
    onlyfilename = onlyfilename.replace(' ','_')
    onlyfilename = onlyfilename.replace('-','_')
    onlyfileext = fullfilename.split('.')[1];
    print(request.files);
    newpath = "uploaded_files/" + onlyfilename  + "_" + datetime.now(pytz.timezone('Asia/Bangkok')).strftime('%Y_%m_%d_%H_%M_%S') + "." + onlyfileext;
    app.logger.info("uploaded new file path : "+newpath)
    file.save(newpath)

    data = {
        "status":"true",
        "upload_quotationfilerecord":
            {
                "result": "pass",
                "full uploaded file path": newpath
            }
    }

    #return json response
    return jsonify(data)

@app.route('/quotation_file_record/download', methods=['POST'])
def get_quotationfilerecord_download():
    app.logger.info('/quotation_file_record/download')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()

    sql = "select * from quotation_file_record where "
    sql += "quot_file_id=";
    sql += request.form.get('quot_file_id')
    cursor.execute(sql)
    results = cursor.fetchall()
    cursor.close()
    conn.close()
    data = {
        "status":"true",
        "show download link":
            {
                "document_name": results[0].quot_name,
                "document_path": results[0].quot_path,
                "document_download_link": "http://deploy-aws.com:3006/downloadfiletocomputer?fileurl="+results[0].quot_name
            }
    }
    return jsonify(data)

@app.route('/news_info/listall', methods=['POST'])
def get_newsinfo_listall():
    app.logger.info('/news_info/listall')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()

    sql = "select cast(news_id as char) as news_id,"
    sql += "cast(title as char) as title,"
    sql += "cast(content as char) as content,"
    sql += "cast(news_date as char) as news_date,"
    sql += "cast(showing_order as char) as showing_order,"
    sql += "cast(short_content as char) as short_content,"
    sql += "cast(short_content as char) as short_content,"
    sql += "cast(news_up_time as char) as news_up_time"
    sql += " from news_info"

    cursor.execute(sql)
    data = cursor.fetchall()
    print(data)
    cursor.close()
    conn.close()
    return jsonify(data)

@app.route('/news_info/update', methods=['POST'])
def get_newsinfo_update():
    app.logger.info('/news_info/update')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()

    sql = "update news_info set "

    if request.form.get('title') is not None:
        sql += ","
        sql += "title='"
        sql += request.form.get('title')
        sql += "'"

    if request.form.get('content') is not None:
        sql += ","
        sql += "content='"
        sql += request.form.get('content')
        sql += "'"

    sql += ",";
    sql += "news_date='";
    sql += datetime.now(pytz.timezone('Asia/Bangkok')).strftime('%Y-%m-%d')
    sql += "'"

    if request.form.get('showing_order') is not None:
        sql += ","
        sql += "showing_order="
        sql += request.form.get('showing_order')

    if request.form.get('short_content') is not None:
        sql += ","
        sql += "short_content='"
        sql += request.form.get('short_content')
        sql += "'"

    sql += " where news_id="
    sql += request.form.get('news_id')

    sql = sql.replace("update news_info set ,", "update news_info set ")

    print('sql='+sql)

    cursor.execute(sql)
    conn.commit()

    cursor.close()
    conn.close()

    data = {
        "status":"true",
        "update news_info":
            {
                "result": "pass"
            }
    }
    return jsonify(data)

@app.route('/news_info/delete', methods=['POST'])
def get_newsinfo_delete():
    app.logger.info('/news_info/delete')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()

    sql = "delete from news_info"
    sql += " where news_id="
    sql += request.form.get('news_id')

    print('sql='+sql)

    cursor.execute(sql)
    conn.commit()

    cursor.close()
    conn.close()

    data = {
        "status":"true",
        "delete news_info":
            {
                "result": "pass"
            }
    }
    return jsonify(data)

@app.route('/news_info/add', methods=['POST'])
def get_newsinfo_add():
    app.logger.info('/news_info/add')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()

    sql = "insert into news_info(title,content,news_date,showing_order,short_content)";
    sql += " values ("

    sql += "'"
    if request.form.get('title') is not None:
        sql += request.form.get('title')
    sql += "'"

    sql += ","

    sql += "'"
    if request.form.get('content') is not None:
        sql += request.form.get('content')
    sql += "'"

    sql += ","

    sql += "'"
    sql += datetime.now(pytz.timezone('Asia/Bangkok')).strftime('%Y-%m-%d')
    sql += "'"

    sql += ","

    if request.form.get('showing_order') is not None:
        sql += request.form.get('showing_order')
    else:
        sql += "null"

    sql += ","

    sql += "'"
    if request.form.get('short_content') is not None:
        sql += request.form.get('short_content')
    sql += "'"

    sql += ")"

    print('sql='+sql)

    cursor.execute(sql)
    conn.commit()

    cursor.close()
    conn.close()

    data = {
        "status":"true",
        "add news_info":
            {
                "result": "pass"
            }
    }
    return jsonify(data)

@app.route('/quotation_list/listall', methods=['POST'])
def get_quotationlist_listall():
    app.logger.info('/quotation_list/listall')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()
    sql = "select * from quotation_list"
    cursor.execute(sql)
    data = cursor.fetchall()
    print(data)
    cursor.close()
    conn.close()
    return jsonify(data)

@app.route('/quotation_list/update', methods=['POST'])
def get_quotationlist_update():
    app.logger.info('/quotation_list/update')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()

    sql = "update quotation_list set "

    if request.form.get('quot_no') is not None:
        sql += ","
        sql += "quot_no='"
        sql += request.form.get('quot_no')
        sql += "'"

    if request.form.get('user_id') is not None:
        sql += ","
        sql += "user_id="
        sql += request.form.get('user_id')

    sql += ",";
    sql += "update_time='";
    sql += datetime.now(pytz.timezone('Asia/Bangkok')).strftime('%Y-%m-%d')
    sql += "'"

    if request.form.get('quot_stat') is not None:
        sql += ","
        sql += "quot_stat='"
        sql += request.form.get('quot_stat')
        sql += "'"

    if request.form.get('quot_ver') is not None:
        sql += ","
        sql += "quot_ver="
        sql += request.form.get('quot_ver')


    sql += " where  quot_id="
    sql += request.form.get(' quot_id')

    sql = sql.replace("update quotation_list set ,", "update quotation_list set ")

    print('sql='+sql)

    cursor.execute(sql)
    conn.commit()

    cursor.close()
    conn.close()

    data = {
        "status":"true",
        "update quotation_list":
            {
                "result": "pass"
            }
    }
    return jsonify(data)

@app.route('/quotation_list/delete', methods=['POST'])
def get_quotationlist_delete():
    app.logger.info('/quotation_list/delete')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()

    sql = "delete quotation_list"
    sql += " where  quot_id="
    sql += request.form.get(' quot_id')

    sql = sql.replace("update quotation_list set ,", "update quotation_list set ")

    print('sql='+sql)

    cursor.execute(sql)
    conn.commit()

    cursor.close()
    conn.close()

    data = {
        "status":"true",
        "delete quotation_list":
            {
                "result": "pass"
            }
    }
    return jsonify(data)

@app.route('/quotation_list/add', methods=['POST'])
def get_quotationlist_add():
    app.logger.info('/quotation_list/add')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()

    sql = "insert into quotation_list(quot_no,user_id,update_time,quot_stat,quot_ver)";
    sql += " values ("

    sql += "'"
    if request.form.get('quot_no') is not None:
        sql += request.form.get('quot_no')
    sql += "'"

    sql += ","


    if request.form.get('user_id') is not None:
        sql += request.form.get('user_id')
    else:
        sql += "null"


    sql += ","

    sql += "'"
    sql += datetime.now(pytz.timezone('Asia/Bangkok')).strftime('%Y-%m-%d')
    sql += "'"

    sql += ","

    sql += "'"
    if request.form.get('quot_stat') is not None:
        sql += request.form.get('quot_stat')
    sql += "'"

    sql += ","


    if request.form.get('quot_ver') is not None:
        sql += request.form.get('quot_ver')
    else:
        sql += "null"

    sql += ")"

    print('sql='+sql)

    cursor.execute(sql)
    conn.commit()

    cursor.close()
    conn.close()

    data = {
        "status":"true",
        "add quotation_list":
            {
                "result": "pass"
            }
    }
    return jsonify(data)

@app.route('/quotation_product/listall', methods=['POST'])
def get_quotation_listall():
    app.logger.info('/quotation_product/listall')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()
    sql = "select * from quotation_product"
    cursor.execute(sql)
    data = cursor.fetchall()
    print(data)
    cursor.close()
    conn.close()
    return jsonify(data)

@app.route('/quotation_product/getquotationbyid', methods=['POST'])
def get_quotation_getquotationbyid():
    app.logger.info('/quotation_product/getquotationbyid')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()
    sql = "select * from quotation_product where quotation_product_id=" + request.form.get('quotation_product_id')
    cursor.execute(sql)
    data = cursor.fetchall()
    print(data)
    cursor.close()
    conn.close()
    return jsonify(data)
@app.route('/quotation_product/update', methods=['POST'])
def get_quotationproduct_update():
    app.logger.info('/quotation_product/update')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()

    sql = "update quotation_product set "

    if request.form.get('quot_id') is not None:
        sql += ","
        sql += "quot_id="
        sql += request.form.get('quot_id')


    if request.form.get('Id') is not None:
        sql += ","
        sql += "Id="
        sql += request.form.get('Id')

    if request.form.get('quantity') is not None:
        sql += ","
        sql += "quantity="
        sql += request.form.get('quantity')

    if request.form.get('unit_price') is not None:
        sql += ","
        sql += "unit_price="
        sql += request.form.get('unit_price')

    if request.form.get('total_price') is not None:
        sql += ","
        sql += "total_price="
        sql += request.form.get('total_price')

    sql += " where quotation_product_id="
    sql += request.form.get('quotation_product_id')

    sql = sql.replace("update quotation_product set ,", "update quotation_product set ")

    print('sql='+sql)

    cursor.execute(sql)
    conn.commit()

    cursor.close()
    conn.close()

    data = {
        "status":"true",
        "update quotation_product":
            {
                "result": "pass"
            }
    }
    return jsonify(data)
@app.route('/quotation_product/delete', methods=['POST'])
def get_quotationproduct_delete():
    app.logger.info('/quotation_product/delete')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()

    sql = "delete from quotation_product"
    sql += " where quotation_product_id="
    sql += request.form.get('quotation_product_id')

    print('sql=' + sql)

    cursor.execute(sql)
    conn.commit()

    cursor.close()
    conn.close()

    data = {
        "status": "true",
        "delete quotation_product":
            {
                "result": "pass"
            }
    }
    return jsonify(data)


@app.route('/quotation_product/add', methods=['POST'])
def get_quotationproduct_add():
    app.logger.info('/quotation_product/add')
    conn = connection_pool.get_connection()
    cursor = conn.cursor()

    sql = "insert into quotation_product(quot_id,Id,quantity,unit_price,total_price) values (";

    if request.form.get('quot_id') is not None:
        sql += request.form.get('quot_id')
    else:
        sql += "null"

    sql += ","

    if request.form.get('Id') is not None:
        sql += request.form.get('Id')
    else:
        sql += "null"

    sql += ","

    if request.form.get('quantity') is not None:
        sql += request.form.get('quantity')
    else:
        sql += "null"

    sql += ","

    if request.form.get('unit_price') is not None:
        sql += request.form.get('unit_price')
    else:
        sql += "null"

    sql += ","

    if request.form.get('total_price') is not None:
        sql += request.form.get('total_price')
    else:
        sql += "null"

    sql += ")"

    print('sql=' + sql)

    cursor.execute(sql)
    conn.commit()

    cursor.close()
    conn.close()

    data = {
        "status": "true",
        "add quotation_product":
            {
                "result": "pass"
            }
    }
    return jsonify(data)




# In-memory data store
# items = [{"id": 1, "name": "This is item 1"}, {"id": 2, "name": "This is item 2"}]

# GET request: Retrieve all items
# @app.route('/api/items', methods=['GET'])
# def get_items():
#     return jsonify(items)

# GET request: Retrieve a specific item by ID
# @app.route('/api/items/<int:item_id>', methods=['GET'])
# def get_item(item_id):
#     item = next((item for item in items if item["id"] == item_id), None)
#     if item is None:
#         return jsonify({"error": "Item not found"}), 404
#     return jsonify(item)

# POST request: Create a new item
# @app.route('/api/items', methods=['POST'])
# def create_item():
#     new_item = {"id": len(items) + 1, "name": request.json.get('name')}
#     items.append(new_item)
#     return jsonify(new_item), 201



if __name__ == "__main__":
    #app.run(debug=True)
    app.run(host='0.0.0.0',debug=True)
